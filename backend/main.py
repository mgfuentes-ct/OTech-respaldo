from fastapi import FastAPI, HTTPException, Form, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from typing import Optional, List
from datetime import datetime, date
from io import BytesIO
import re
import logging

from passlib.context import CryptContext
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import get_db_connection


# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================

app = FastAPI(title="OTech - MXD Inventory API")

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


# ============================================================
# SCHEMAS
# ============================================================

class BuscarCodigoRequest(BaseModel):
    codigo: str


class CrearProductoRequest(BaseModel):
    codigo_original: str
    nombre: str
    descripcion: Optional[str] = None
    stock_minimo: int = 0
    id_marca: int
    id_tipo_producto: int
    id_modelo: Optional[int] = None
    modelos_compatibles: Optional[List[int]] = []


class RegistrarInventarioRequest(BaseModel):
    id_producto: int
    numero_serie: str
    fecha_entrada: Optional[date] = None
    id_estado: int
    id_ubicacion: Optional[int] = None
    id_usuario_registro: int
    vendido: bool = False
    prestamo: bool = False
    para_venta: bool = False
    destinatario: Optional[str] = None
    observaciones: Optional[str] = None


class ActualizarInventarioRequest(BaseModel):
    id_inventario: int
    id_estado: int
    id_ubicacion: Optional[int] = None
    fecha_salida: Optional[date] = None
    vendido: bool = False
    prestamo: bool = False
    para_venta: bool = False
    destinatario: Optional[str] = None
    observaciones: Optional[str] = None
    id_usuario: int


# Compatibilidad con tu frontend anterior.
# Si todavía mandas id_pieza/nuevo_estado/caja desde renderer.js,
# este schema permite que el endpoint /actualizar_estado_pieza siga funcionando.
class ActualizarEstadoCompatRequest(BaseModel):
    id_pieza: Optional[int] = None
    id_inventario: Optional[int] = None
    nuevo_estado: Optional[str] = None
    id_estado: Optional[int] = None
    caja: Optional[str] = None
    id_ubicacion: Optional[int] = None
    id_usuario: int
    observaciones: Optional[str] = None


class CrearUsuarioRequest(BaseModel):
    nombre_completo: str
    nombre_usuario: str
    email: str
    password: str


# ============================================================
# HELPERS
# ============================================================

def cerrar_conexion(cursor=None, conn=None):
    try:
        if cursor:
            cursor.close()
    except Exception:
        pass

    try:
        if conn and conn.is_connected():
            conn.close()
    except Exception:
        pass


def obtener_id_estado_por_nombre(cursor, nombre_estado: str):
    cursor.execute(
        "SELECT id_estado FROM estado WHERE nombre = %s",
        (nombre_estado,)
    )
    row = cursor.fetchone()
    return row["id_estado"] if row else None


def obtener_nombre_estado_por_id(cursor, id_estado: int):
    cursor.execute(
        "SELECT nombre FROM estado WHERE id_estado = %s",
        (id_estado,)
    )
    row = cursor.fetchone()
    return row["nombre"] if row else None


def obtener_codigo_ubicacion_por_id(cursor, id_ubicacion: Optional[int]):
    if not id_ubicacion:
        return None

    cursor.execute(
        "SELECT codigo FROM ubicacion WHERE id_ubicacion = %s",
        (id_ubicacion,)
    )
    row = cursor.fetchone()
    return row["codigo"] if row else None


def registrar_movimiento_inventario(
    conn,
    cursor,
    id_inventario: int,
    tipo_movimiento: str,
    id_usuario: int,
    estado_anterior: Optional[str] = None,
    estado_nuevo: Optional[str] = None,
    ubicacion_anterior: Optional[str] = None,
    ubicacion_nueva: Optional[str] = None,
    destinatario_anterior: Optional[str] = None,
    destinatario_nuevo: Optional[str] = None,
    observaciones: Optional[str] = None
):
    cursor.execute("""
        INSERT INTO movimiento (
            id_inventario,
            tipo_movimiento,
            estado_anterior,
            estado_nuevo,
            ubicacion_anterior,
            ubicacion_nueva,
            destinatario_anterior,
            destinatario_nuevo,
            observaciones,
            id_usuario
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        id_inventario,
        tipo_movimiento,
        estado_anterior,
        estado_nuevo,
        ubicacion_anterior,
        ubicacion_nueva,
        destinatario_anterior,
        destinatario_nuevo,
        observaciones,
        id_usuario
    ))


# ============================================================
# LOGIN
# ============================================================

@app.post("/login")
async def login(username: str = Form(...), password: str = Form(...)):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            "SELECT * FROM usuario WHERE nombre_usuario = %s",
            (username,)
        )
        user = cursor.fetchone()

        if not user:
            raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")

        if not user["activo"]:
            raise HTTPException(status_code=401, detail="Usuario inactivo. Contacte al administrador.")

        if not pwd_context.verify(password, user["password_hash"]):
            raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")

        cursor.execute(
            "UPDATE usuario SET ultimo_login = NOW() WHERE id_usuario = %s",
            (user["id_usuario"],)
        )

        cursor.execute("""
            INSERT INTO log_sesion (id_usuario, fecha_login)
            VALUES (%s, NOW())
        """, (user["id_usuario"],))

        conn.commit()

        return {
            "id_usuario": user["id_usuario"],
            "nombre_usuario": user["nombre_usuario"],
            "nombre_completo": user["nombre_completo"],
            "rol": user["rol"],
            "token": "mock-token-" + str(user["id_usuario"])
        }

    finally:
        cerrar_conexion(cursor, conn)


# ============================================================
# CATÁLOGOS
# ============================================================

@app.get("/catalogos/marcas")
def listar_marcas():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("SELECT id_marca, nombre, descripcion FROM marca ORDER BY nombre")
        return cursor.fetchall()
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/catalogos/modelos")
def listar_modelos(id_marca: Optional[int] = Query(None)):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        if id_marca:
            cursor.execute("""
                SELECT 
                    mo.id_modelo,
                    mo.nombre,
                    mo.id_marca,
                    ma.nombre AS marca
                FROM modelo mo
                JOIN marca ma ON mo.id_marca = ma.id_marca
                WHERE mo.id_marca = %s
                ORDER BY mo.nombre
            """, (id_marca,))
        else:
            cursor.execute("""
                SELECT 
                    mo.id_modelo,
                    mo.nombre,
                    mo.id_marca,
                    ma.nombre AS marca
                FROM modelo mo
                JOIN marca ma ON mo.id_marca = ma.id_marca
                ORDER BY ma.nombre, mo.nombre
            """)

        return cursor.fetchall()
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/catalogos/tipos_producto")
def listar_tipos_producto():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("SELECT id_tipo_producto, nombre FROM tipo_producto ORDER BY nombre")
        return cursor.fetchall()
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/catalogos/estados")
def listar_estados():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("SELECT id_estado, nombre FROM estado ORDER BY nombre")
        return cursor.fetchall()
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/catalogos/ubicaciones")
def listar_ubicaciones(id_marca: Optional[int] = Query(None)):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        if id_marca:
            cursor.execute("""
                SELECT 
                    u.id_ubicacion,
                    u.codigo,
                    u.descripcion,
                    a.id_anaquel,
                    a.nombre AS anaquel,
                    ma.id_marca,
                    ma.nombre AS marca
                FROM ubicacion u
                JOIN anaquel a ON u.id_anaquel = a.id_anaquel
                JOIN marca ma ON a.id_marca = ma.id_marca
                WHERE ma.id_marca = %s
                ORDER BY a.nombre, u.codigo
            """, (id_marca,))
        else:
            cursor.execute("""
                SELECT 
                    u.id_ubicacion,
                    u.codigo,
                    u.descripcion,
                    a.id_anaquel,
                    a.nombre AS anaquel,
                    ma.id_marca,
                    ma.nombre AS marca
                FROM ubicacion u
                JOIN anaquel a ON u.id_anaquel = a.id_anaquel
                JOIN marca ma ON a.id_marca = ma.id_marca
                ORDER BY ma.nombre, a.nombre, u.codigo
            """)

        return cursor.fetchall()
    finally:
        cerrar_conexion(cursor, conn)


# Compatibilidad con endpoint anterior
@app.get("/admin/listar_drones")
async def listar_drones_compat():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT 
                mo.id_modelo AS id,
                CONCAT(ma.nombre, ' ', mo.nombre) AS nombre,
                mo.id_modelo,
                ma.id_marca,
                ma.nombre AS marca
            FROM modelo mo
            JOIN marca ma ON mo.id_marca = ma.id_marca
            ORDER BY ma.nombre, mo.nombre
        """)
        return cursor.fetchall()
    finally:
        cerrar_conexion(cursor, conn)


# ============================================================
# PRODUCTOS
# ============================================================

@app.get("/productos")
def listar_productos():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT 
                p.id_producto,
                p.codigo_original,
                p.nombre,
                p.descripcion,
                p.stock_minimo,
                ma.id_marca,
                ma.nombre AS marca,
                tp.id_tipo_producto,
                tp.nombre AS tipo_producto,
                mo.id_modelo,
                mo.nombre AS modelo_base,
                GROUP_CONCAT(moc.nombre ORDER BY moc.nombre SEPARATOR ', ') AS modelos_compatibles
            FROM producto p
            JOIN marca ma ON p.id_marca = ma.id_marca
            JOIN tipo_producto tp ON p.id_tipo_producto = tp.id_tipo_producto
            LEFT JOIN modelo mo ON p.id_modelo = mo.id_modelo
            LEFT JOIN producto_modelo_compatible pmc ON p.id_producto = pmc.id_producto
            LEFT JOIN modelo moc ON pmc.id_modelo = moc.id_modelo
            GROUP BY p.id_producto
            ORDER BY ma.nombre, tp.nombre, p.nombre
        """)

        return cursor.fetchall()
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/admin/crear_producto")
async def crear_producto_admin(data: CrearProductoRequest):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            "SELECT id_producto FROM producto WHERE codigo_original = %s",
            (data.codigo_original,)
        )
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail="El código original ya está registrado")

        cursor.execute("""
            INSERT INTO producto (
                id_marca,
                id_tipo_producto,
                id_modelo,
                codigo_original,
                nombre,
                descripcion,
                stock_minimo
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            data.id_marca,
            data.id_tipo_producto,
            data.id_modelo,
            data.codigo_original.strip().upper(),
            data.nombre.strip(),
            data.descripcion,
            data.stock_minimo
        ))

        id_producto = cursor.lastrowid

        for id_modelo in data.modelos_compatibles or []:
            cursor.execute("""
                INSERT INTO producto_modelo_compatible (id_producto, id_modelo)
                VALUES (%s, %s)
            """, (id_producto, id_modelo))

        conn.commit()

        return {
            "mensaje": f"Producto '{data.nombre}' creado exitosamente",
            "id_producto": id_producto
        }

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error al crear producto: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cerrar_conexion(cursor, conn)


# ============================================================
# BÚSQUEDA POR CÓDIGO
# ============================================================

@app.post("/buscar_codigo")
async def buscar_codigo_endpoint(data: BuscarCodigoRequest):
    codigo = data.codigo.strip().upper()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # 1. Buscar primero como número de serie
        cursor.execute("""
            SELECT 
                i.id_inventario,
                i.numero_serie,
                i.fecha_entrada,
                i.fecha_salida,
                i.vendido,
                i.prestamo,
                i.para_venta,
                i.destinatario,
                i.observaciones,
                i.id_estado,
                i.id_ubicacion,
                p.id_producto,
                p.codigo_original,
                p.nombre AS nombre_producto,
                ma.nombre AS marca,
                tp.nombre AS tipo_producto,
                mo.nombre AS modelo,
                e.nombre AS estado,
                u.codigo AS ubicacion,
                a.nombre AS anaquel
            FROM inventario i
            JOIN producto p ON i.id_producto = p.id_producto
            JOIN marca ma ON p.id_marca = ma.id_marca
            JOIN tipo_producto tp ON p.id_tipo_producto = tp.id_tipo_producto
            LEFT JOIN modelo mo ON p.id_modelo = mo.id_modelo
            LEFT JOIN estado e ON i.id_estado = e.id_estado
            LEFT JOIN ubicacion u ON i.id_ubicacion = u.id_ubicacion
            LEFT JOIN anaquel a ON u.id_anaquel = a.id_anaquel
            WHERE UPPER(i.numero_serie) = %s
              AND i.activo = 1
        """, (codigo,))
        inventario = cursor.fetchone()

        if inventario:
            return {
                "tipo": "numero_serie",
                "inventario": inventario,
                # compatibilidad con renderer anterior
                "pieza": {
                    "id_pieza": inventario["id_inventario"],
                    "id_inventario": inventario["id_inventario"],
                    "numero_serie": inventario["numero_serie"],
                    "estado": inventario["estado"],
                    "caja": inventario["ubicacion"],
                    "nombre_producto": inventario["nombre_producto"]
                }
            }

        # 2. Buscar como número de parte / código original
        cursor.execute("""
            SELECT 
                p.id_producto,
                p.codigo_original,
                p.nombre,
                p.descripcion,
                p.stock_minimo,
                p.id_marca,
                p.id_tipo_producto,
                p.id_modelo,
                ma.nombre AS marca,
                tp.nombre AS tipo_producto,
                mo.nombre AS modelo,
                GROUP_CONCAT(moc.nombre ORDER BY moc.nombre SEPARATOR ', ') AS modelos_compatibles
            FROM producto p
            JOIN marca ma ON p.id_marca = ma.id_marca
            JOIN tipo_producto tp ON p.id_tipo_producto = tp.id_tipo_producto
            LEFT JOIN modelo mo ON p.id_modelo = mo.id_modelo
            LEFT JOIN producto_modelo_compatible pmc ON p.id_producto = pmc.id_producto
            LEFT JOIN modelo moc ON pmc.id_modelo = moc.id_modelo
            WHERE UPPER(p.codigo_original) = %s
            GROUP BY p.id_producto
        """, (codigo,))
        producto = cursor.fetchone()

        if producto:
            cursor.execute("""
                SELECT 
                    i.id_inventario,
                    i.numero_serie,
                    e.nombre AS estado,
                    u.codigo AS ubicacion,
                    i.destinatario,
                    i.para_venta,
                    i.vendido,
                    i.prestamo
                FROM inventario i
                LEFT JOIN estado e ON i.id_estado = e.id_estado
                LEFT JOIN ubicacion u ON i.id_ubicacion = u.id_ubicacion
                WHERE i.id_producto = %s
                  AND i.activo = 1
                ORDER BY i.fecha_registro DESC
            """, (producto["id_producto"],))
            items = cursor.fetchall()

            return {
                "tipo": "numero_parte",
                "producto": producto,
                "inventario": items,
                # compatibilidad con renderer anterior
                "piezas": items
            }

        # 3. No encontrado
        return {
            "tipo": "no_encontrado",
            "codigo_original": codigo,
            "mensaje": "No existe ningún producto o pieza con ese número de parte o número de serie."
        }

    except Exception as e:
        logger.error(f"Error en /buscar_codigo: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cerrar_conexion(cursor, conn)


# ============================================================
# INVENTARIO
# ============================================================

@app.get("/inventario")
async def obtener_inventario():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT 
                i.id_inventario,
                i.id_inventario AS id_pieza,
                i.numero_serie,
                i.fecha_entrada,
                i.fecha_salida,
                i.vendido,
                i.prestamo,
                i.para_venta,
                i.destinatario,
                i.observaciones,
                i.fecha_registro,
                i.activo,
                p.id_producto,
                p.codigo_original,
                p.nombre AS producto,
                p.nombre AS nombre_producto,
                ma.nombre AS marca,
                tp.nombre AS tipo_producto,
                mo.nombre AS modelo,
                e.id_estado,
                e.nombre AS estado,
                u.id_ubicacion,
                u.codigo AS ubicacion,
                u.codigo AS caja,
                a.nombre AS anaquel,
                COALESCE(us.nombre_usuario, 'Usuario eliminado') AS usuario_registro,
                COALESCE(us.nombre_usuario, 'Usuario eliminado') AS nombre_usuario,
                GROUP_CONCAT(DISTINCT moc.nombre ORDER BY moc.nombre SEPARATOR ', ') AS modelos_compatibles
            FROM inventario i
            JOIN producto p ON i.id_producto = p.id_producto
            JOIN marca ma ON p.id_marca = ma.id_marca
            JOIN tipo_producto tp ON p.id_tipo_producto = tp.id_tipo_producto
            LEFT JOIN modelo mo ON p.id_modelo = mo.id_modelo
            LEFT JOIN estado e ON i.id_estado = e.id_estado
            LEFT JOIN ubicacion u ON i.id_ubicacion = u.id_ubicacion
            LEFT JOIN anaquel a ON u.id_anaquel = a.id_anaquel
            LEFT JOIN usuario us ON i.id_usuario_registro = us.id_usuario
            LEFT JOIN producto_modelo_compatible pmc ON p.id_producto = pmc.id_producto
            LEFT JOIN modelo moc ON pmc.id_modelo = moc.id_modelo
            WHERE i.activo = 1
            GROUP BY i.id_inventario
            ORDER BY i.fecha_registro DESC
        """)

        return cursor.fetchall()

    except Exception as e:
        logger.error(f"Error en /inventario: {e}")
        raise HTTPException(status_code=500, detail=f"Error en consulta SQL: {str(e)}")
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/inventario/registrar")
async def registrar_inventario(data: RegistrarInventarioRequest):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            "SELECT id_inventario FROM inventario WHERE numero_serie = %s",
            (data.numero_serie.strip().upper(),)
        )
        if cursor.fetchone():
            raise HTTPException(status_code=400, detail="El número de serie ya está registrado")

        estado_nuevo = obtener_nombre_estado_por_id(cursor, data.id_estado)
        ubicacion_nueva = obtener_codigo_ubicacion_por_id(cursor, data.id_ubicacion)

        cursor.execute("""
            INSERT INTO inventario (
                id_producto,
                numero_serie,
                fecha_entrada,
                fecha_salida,
                vendido,
                prestamo,
                para_venta,
                destinatario,
                id_estado,
                observaciones,
                id_ubicacion,
                id_usuario_registro,
                activo
            )
            VALUES (%s, %s, COALESCE(%s, CURDATE()), %s, %s, %s, %s, %s, %s, %s, %s, %s, 1)
        """, (
            data.id_producto,
            data.numero_serie.strip().upper(),
            data.fecha_entrada,
            None,
            data.vendido,
            data.prestamo,
            data.para_venta,
            data.destinatario,
            data.id_estado,
            data.observaciones,
            data.id_ubicacion,
            data.id_usuario_registro
        ))

        id_inventario = cursor.lastrowid

        registrar_movimiento_inventario(
            conn=conn,
            cursor=cursor,
            id_inventario=id_inventario,
            tipo_movimiento="registro",
            id_usuario=data.id_usuario_registro,
            estado_nuevo=estado_nuevo,
            ubicacion_nueva=ubicacion_nueva,
            destinatario_nuevo=data.destinatario,
            observaciones=data.observaciones or "Inventario registrado"
        )

        cursor.execute("""
            SELECT
                i.id_inventario,
                i.numero_serie,
                i.fecha_entrada,
                p.codigo_original,
                p.nombre AS nombre_producto,
                ma.nombre AS marca,
                tp.nombre AS tipo_producto,
                mo.nombre AS modelo,
                e.nombre AS estado,
                u.codigo AS ubicacion
            FROM inventario i
            JOIN producto p ON i.id_producto = p.id_producto
            JOIN marca ma ON p.id_marca = ma.id_marca
            JOIN tipo_producto tp ON p.id_tipo_producto = tp.id_tipo_producto
            LEFT JOIN modelo mo ON p.id_modelo = mo.id_modelo
            LEFT JOIN estado e ON i.id_estado = e.id_estado
            LEFT JOIN ubicacion u ON i.id_ubicacion = u.id_ubicacion
            WHERE i.id_inventario = %s
        """, (id_inventario,))
        inventario_creado = cursor.fetchone()

        conn.commit()

        return {
            "mensaje": "Inventario registrado correctamente",
            "id_inventario": id_inventario,
            "inventario": inventario_creado
        }

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error al registrar inventario: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/inventario/actualizar")
async def actualizar_inventario(data: ActualizarInventarioRequest):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT 
                i.id_estado,
                e.nombre AS estado,
                i.id_ubicacion,
                u.codigo AS ubicacion,
                i.destinatario
            FROM inventario i
            LEFT JOIN estado e ON i.id_estado = e.id_estado
            LEFT JOIN ubicacion u ON i.id_ubicacion = u.id_ubicacion
            WHERE i.id_inventario = %s
        """, (data.id_inventario,))
        anterior = cursor.fetchone()

        if not anterior:
            raise HTTPException(status_code=404, detail="Inventario no encontrado")

        estado_nuevo = obtener_nombre_estado_por_id(cursor, data.id_estado)
        ubicacion_nueva = obtener_codigo_ubicacion_por_id(cursor, data.id_ubicacion)

        if not estado_nuevo:
            raise HTTPException(status_code=400, detail="Estado no válido")

        tipo_movimiento = "cambio_estado"

        if data.vendido:
            tipo_movimiento = "venta"
        elif data.prestamo:
            tipo_movimiento = "prestamo"
        elif data.id_ubicacion != anterior["id_ubicacion"]:
            tipo_movimiento = "cambio_ubicacion"

        cursor.execute("""
            UPDATE inventario
            SET 
                id_estado = %s,
                id_ubicacion = %s,
                fecha_salida = %s,
                vendido = %s,
                prestamo = %s,
                para_venta = %s,
                destinatario = %s,
                observaciones = %s
            WHERE id_inventario = %s
        """, (
            data.id_estado,
            data.id_ubicacion,
            data.fecha_salida,
            data.vendido,
            data.prestamo,
            data.para_venta,
            data.destinatario,
            data.observaciones,
            data.id_inventario
        ))

        registrar_movimiento_inventario(
            conn=conn,
            cursor=cursor,
            id_inventario=data.id_inventario,
            tipo_movimiento=tipo_movimiento,
            id_usuario=data.id_usuario,
            estado_anterior=anterior["estado"],
            estado_nuevo=estado_nuevo,
            ubicacion_anterior=anterior["ubicacion"],
            ubicacion_nueva=ubicacion_nueva,
            destinatario_anterior=anterior["destinatario"],
            destinatario_nuevo=data.destinatario,
            observaciones=data.observaciones
        )

        conn.commit()

        return {"mensaje": "Inventario actualizado correctamente"}

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error al actualizar inventario: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cerrar_conexion(cursor, conn)


# Compatibilidad con tu endpoint anterior /actualizar_estado_pieza
@app.post("/actualizar_estado_pieza")
async def actualizar_estado_pieza_endpoint(data: ActualizarEstadoCompatRequest):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        id_inventario = data.id_inventario or data.id_pieza

        if not id_inventario:
            raise HTTPException(status_code=400, detail="Falta id_inventario")

        if data.id_estado:
            id_estado_nuevo = data.id_estado
        elif data.nuevo_estado:
            id_estado_nuevo = obtener_id_estado_por_nombre(cursor, data.nuevo_estado)
        else:
            raise HTTPException(status_code=400, detail="Falta estado nuevo")

        if not id_estado_nuevo:
            raise HTTPException(status_code=400, detail="Estado no válido")

        cursor.execute("""
            SELECT 
                i.id_estado,
                e.nombre AS estado,
                i.id_ubicacion,
                u.codigo AS ubicacion,
                i.destinatario
            FROM inventario i
            LEFT JOIN estado e ON i.id_estado = e.id_estado
            LEFT JOIN ubicacion u ON i.id_ubicacion = u.id_ubicacion
            WHERE i.id_inventario = %s
        """, (id_inventario,))
        anterior = cursor.fetchone()

        if not anterior:
            raise HTTPException(status_code=404, detail="Inventario no encontrado")

        id_ubicacion_nueva = data.id_ubicacion

        # Si todavía mandas "caja" como texto, se intenta encontrar como código de ubicación.
        if not id_ubicacion_nueva and data.caja:
            cursor.execute(
                "SELECT id_ubicacion FROM ubicacion WHERE codigo = %s",
                (data.caja,)
            )
            ubicacion = cursor.fetchone()

            if ubicacion:
                id_ubicacion_nueva = ubicacion["id_ubicacion"]
            else:
                raise HTTPException(
                    status_code=400,
                    detail="La ubicación/caja no existe. Regístrala en la tabla ubicacion."
                )

        if not id_ubicacion_nueva:
            id_ubicacion_nueva = anterior["id_ubicacion"]

        estado_nuevo = obtener_nombre_estado_por_id(cursor, id_estado_nuevo)
        ubicacion_nueva = obtener_codigo_ubicacion_por_id(cursor, id_ubicacion_nueva)

        cursor.execute("""
            UPDATE inventario
            SET 
                id_estado = %s,
                id_ubicacion = %s,
                observaciones = %s
            WHERE id_inventario = %s
        """, (
            id_estado_nuevo,
            id_ubicacion_nueva,
            data.observaciones,
            id_inventario
        ))

        registrar_movimiento_inventario(
            conn=conn,
            cursor=cursor,
            id_inventario=id_inventario,
            tipo_movimiento="cambio_estado",
            id_usuario=data.id_usuario,
            estado_anterior=anterior["estado"],
            estado_nuevo=estado_nuevo,
            ubicacion_anterior=anterior["ubicacion"],
            ubicacion_nueva=ubicacion_nueva,
            destinatario_anterior=anterior["destinatario"],
            destinatario_nuevo=anterior["destinatario"],
            observaciones=data.observaciones
        )

        conn.commit()

        return {"mensaje": "Estado y ubicación actualizados correctamente"}

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        logger.error(f"Error en /actualizar_estado_pieza: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cerrar_conexion(cursor, conn)


# ============================================================
# HISTORIAL
# ============================================================

@app.get("/historial/inventario/{id_inventario}")
def obtener_historial_inventario(id_inventario: int):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT 
                m.id_movimiento,
                m.tipo_movimiento,
                m.estado_anterior,
                m.estado_nuevo,
                m.ubicacion_anterior,
                m.ubicacion_nueva,
                m.destinatario_anterior,
                m.destinatario_nuevo,
                m.observaciones,
                m.fecha_movimiento,
                COALESCE(u.nombre_usuario, 'Usuario eliminado') AS nombre_usuario
            FROM movimiento m
            LEFT JOIN usuario u ON m.id_usuario = u.id_usuario
            WHERE m.id_inventario = %s
            ORDER BY m.fecha_movimiento DESC
        """, (id_inventario,))

        return cursor.fetchall()
    finally:
        cerrar_conexion(cursor, conn)


# Compatibilidad con ruta anterior
@app.get("/historial/pieza/{id_pieza}")
def obtener_historial_pieza(id_pieza: int):
    return obtener_historial_inventario(id_pieza)


@app.get("/exportar/historial")
def exportar_historial(
    fecha_inicio: str = Query(...),
    fecha_fin: str = Query(...),
    id_usuario: int = Query(...)
):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            "SELECT rol FROM usuario WHERE id_usuario = %s",
            (id_usuario,)
        )
        usuario = cursor.fetchone()

        if not usuario or usuario["rol"] != "Admin":
            raise HTTPException(status_code=403, detail="Solo el administrador puede exportar historial")

        try:
            inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fin = datetime.strptime(fecha_fin, "%Y-%m-%d")
            fin = fin.replace(hour=23, minute=59, second=59)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha inválido")

        cursor.execute("""
            SELECT 
                m.id_movimiento,
                ma.nombre AS marca,
                tp.nombre AS tipo_producto,
                pr.nombre AS producto,
                mo.nombre AS modelo,
                i.numero_serie,
                m.tipo_movimiento,
                m.estado_anterior,
                m.estado_nuevo,
                m.ubicacion_anterior,
                m.ubicacion_nueva,
                m.destinatario_anterior,
                m.destinatario_nuevo,
                m.observaciones,
                m.fecha_movimiento,
                COALESCE(u.nombre_usuario, 'Usuario eliminado') AS usuario
            FROM movimiento m
            JOIN inventario i ON m.id_inventario = i.id_inventario
            JOIN producto pr ON i.id_producto = pr.id_producto
            JOIN marca ma ON pr.id_marca = ma.id_marca
            JOIN tipo_producto tp ON pr.id_tipo_producto = tp.id_tipo_producto
            LEFT JOIN modelo mo ON pr.id_modelo = mo.id_modelo
            LEFT JOIN usuario u ON m.id_usuario = u.id_usuario
            WHERE m.fecha_movimiento BETWEEN %s AND %s
            ORDER BY m.fecha_movimiento ASC
        """, (inicio, fin))

        movimientos = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"

        headers = [
            "ID Movimiento",
            "Marca",
            "Tipo Producto",
            "Producto",
            "Modelo",
            "Número de Serie",
            "Tipo Movimiento",
            "Estado Anterior",
            "Estado Nuevo",
            "Ubicación Anterior",
            "Ubicación Nueva",
            "Destinatario Anterior",
            "Destinatario Nuevo",
            "Observaciones",
            "Fecha Movimiento",
            "Usuario"
        ]

        header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, mov in enumerate(movimientos, 2):
            values = [
                mov["id_movimiento"],
                mov["marca"],
                mov["tipo_producto"],
                mov["producto"],
                mov["modelo"],
                mov["numero_serie"],
                mov["tipo_movimiento"],
                mov["estado_anterior"],
                mov["estado_nuevo"],
                mov["ubicacion_anterior"],
                mov["ubicacion_nueva"],
                mov["destinatario_anterior"],
                mov["destinatario_nuevo"],
                mov["observaciones"],
                mov["fecha_movimiento"].strftime("%d/%m/%Y %H:%M") if mov["fecha_movimiento"] else "",
                mov["usuario"]
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                max_length = max(max_length, len(str(cell.value or "")))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

        ws.sheet_view.showGridLines = False

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"historial_movimientos_{inicio.strftime('%Y%m%d')}_{fin.strftime('%Y%m%d')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cerrar_conexion(cursor, conn)


# ============================================================
# ALERTAS Y EXPORTACIÓN INVENTARIO
# ============================================================

@app.get("/alertas/stock_bajo")
async def obtener_alertas_stock_bajo():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT 
                p.id_producto,
                p.nombre,
                p.stock_minimo,
                ma.nombre AS marca,
                tp.nombre AS tipo_producto,
                COUNT(i.id_inventario) AS stock_actual
            FROM producto p
            JOIN marca ma ON p.id_marca = ma.id_marca
            JOIN tipo_producto tp ON p.id_tipo_producto = tp.id_tipo_producto
            LEFT JOIN inventario i 
                ON p.id_producto = i.id_producto 
                AND i.activo = 1
            WHERE p.stock_minimo > 0
            GROUP BY p.id_producto
            HAVING stock_actual < p.stock_minimo
            ORDER BY ma.nombre, p.nombre
        """)

        return cursor.fetchall()
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/exportar/inventario")
async def exportar_inventario():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT 
                i.id_inventario,
                ma.nombre AS marca,
                tp.nombre AS tipo_producto,
                pr.nombre AS producto,
                mo.nombre AS modelo,
                i.numero_serie,
                i.fecha_entrada,
                i.fecha_salida,
                i.vendido,
                i.prestamo,
                i.para_venta,
                i.destinatario,
                e.nombre AS estado,
                a.nombre AS anaquel,
                u.codigo AS ubicacion,
                i.observaciones,
                COALESCE(us.nombre_usuario, 'Usuario eliminado') AS usuario_registro,
                i.fecha_registro
            FROM inventario i
            JOIN producto pr ON i.id_producto = pr.id_producto
            JOIN marca ma ON pr.id_marca = ma.id_marca
            JOIN tipo_producto tp ON pr.id_tipo_producto = tp.id_tipo_producto
            LEFT JOIN modelo mo ON pr.id_modelo = mo.id_modelo
            LEFT JOIN estado e ON i.id_estado = e.id_estado
            LEFT JOIN ubicacion u ON i.id_ubicacion = u.id_ubicacion
            LEFT JOIN anaquel a ON u.id_anaquel = a.id_anaquel
            LEFT JOIN usuario us ON i.id_usuario_registro = us.id_usuario
            WHERE i.activo = 1
            ORDER BY ma.nombre, tp.nombre, pr.nombre, i.fecha_registro DESC
        """)

        items = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        headers = [
            "ID Inventario",
            "Marca",
            "Tipo Producto",
            "Producto",
            "Modelo",
            "Número de Serie",
            "Fecha Entrada",
            "Fecha Salida",
            "Vendido",
            "Préstamo",
            "Para Venta",
            "Destinatario",
            "Estado",
            "Anaquel",
            "Ubicación",
            "Observaciones",
            "Registrado Por",
            "Fecha Registro"
        ]

        header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, item in enumerate(items, 2):
            values = [
                item["id_inventario"],
                item["marca"],
                item["tipo_producto"],
                item["producto"],
                item["modelo"],
                item["numero_serie"],
                item["fecha_entrada"].strftime("%d/%m/%Y") if item["fecha_entrada"] else "",
                item["fecha_salida"].strftime("%d/%m/%Y") if item["fecha_salida"] else "",
                "Sí" if item["vendido"] else "No",
                "Sí" if item["prestamo"] else "No",
                "Sí" if item["para_venta"] else "No",
                item["destinatario"],
                item["estado"],
                item["anaquel"],
                item["ubicacion"],
                item["observaciones"],
                item["usuario_registro"],
                item["fecha_registro"].strftime("%d/%m/%Y %H:%M") if item["fecha_registro"] else ""
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                max_length = max(max_length, len(str(cell.value or "")))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

        ws.sheet_view.showGridLines = False

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"inventario_otech_mxd_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        cerrar_conexion(cursor, conn)


# ============================================================
# USUARIOS ADMIN
# ============================================================

@app.get("/admin/listar_usuarios")
async def listar_usuarios():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT 
                id_usuario,
                nombre_usuario,
                nombre_completo,
                email,
                rol,
                activo,
                ultimo_login
            FROM usuario
            ORDER BY id_usuario DESC
        """)
        return cursor.fetchall()
    finally:
        cerrar_conexion(cursor, conn)


@app.post("/admin/crear_usuario")
async def crear_usuario_admin(
    nombre_completo: str = Form(...),
    nombre_usuario: str = Form(...),
    email: str = Form(...),
    password: str = Form(...)
):
    email_regex = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"

    if not re.match(email_regex, email):
        raise HTTPException(status_code=400, detail="Formato de correo electrónico inválido")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT id_usuario
            FROM usuario
            WHERE nombre_usuario = %s
               OR email = %s
               OR nombre_completo = %s
        """, (nombre_usuario, email, nombre_completo))

        if cursor.fetchone():
            raise HTTPException(
                status_code=400,
                detail="El nombre de usuario, correo o nombre completo ya están registrados"
            )

        password_hash = pwd_context.hash(password)

        cursor.execute("""
            INSERT INTO usuario (
                nombre_usuario,
                nombre_completo,
                email,
                rol,
                activo,
                password_hash
            )
            VALUES (%s, %s, %s, 'Operador', 1, %s)
        """, (
            nombre_usuario,
            nombre_completo,
            email,
            password_hash
        ))

        conn.commit()
        user_id = cursor.lastrowid

        return {
            "mensaje": f"Usuario '{nombre_completo}' creado exitosamente con rol 'Operador' (ID {user_id})"
        }

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cerrar_conexion(cursor, conn)


@app.get("/admin/obtener_usuario/{id_usuario}")
async def obtener_usuario(id_usuario: int):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            "SELECT * FROM usuario WHERE id_usuario = %s",
            (id_usuario,)
        )
        usuario = cursor.fetchone()

        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        return usuario
    finally:
        cerrar_conexion(cursor, conn)


@app.put("/admin/editar_usuario/{id_usuario}")
async def editar_usuario(
    id_usuario: int,
    nombre_completo: Optional[str] = None,
    nombre_usuario: Optional[str] = None,
    email: Optional[str] = None,
    rol: Optional[str] = None
):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            "SELECT * FROM usuario WHERE id_usuario = %s",
            (id_usuario,)
        )
        usuario_existente = cursor.fetchone()

        if not usuario_existente:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        if rol and rol not in ["Admin", "Operador"]:
            raise HTTPException(status_code=400, detail="Rol no válido. Solo Admin u Operador")

        updates = []
        valores = []

        if nombre_completo is not None:
            updates.append("nombre_completo = %s")
            valores.append(nombre_completo)

        if nombre_usuario is not None:
            updates.append("nombre_usuario = %s")
            valores.append(nombre_usuario)

        if email is not None:
            updates.append("email = %s")
            valores.append(email)

        if rol is not None:
            updates.append("rol = %s")
            valores.append(rol)

        if not updates:
            raise HTTPException(status_code=400, detail="No se proporcionaron campos para actualizar")

        valores.append(id_usuario)

        query = f"UPDATE usuario SET {', '.join(updates)} WHERE id_usuario = %s"
        cursor.execute(query, valores)

        conn.commit()

        return {"mensaje": f"Usuario ID {id_usuario} actualizado exitosamente"}

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cerrar_conexion(cursor, conn)


@app.put("/admin/eliminar_usuario/{id_usuario}")
async def eliminar_usuario(id_usuario: int):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            "SELECT activo FROM usuario WHERE id_usuario = %s",
            (id_usuario,)
        )
        usuario = cursor.fetchone()

        if not usuario:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        nuevo_estado = 0 if usuario["activo"] else 1

        cursor.execute(
            "UPDATE usuario SET activo = %s WHERE id_usuario = %s",
            (nuevo_estado, id_usuario)
        )

        conn.commit()

        estado_texto = "activado" if nuevo_estado else "desactivado"

        return {"mensaje": f"Usuario ID {id_usuario} {estado_texto} exitosamente"}

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cerrar_conexion(cursor, conn)


# ============================================================
# HEALTH CHECK
# ============================================================

@app.get("/health")
def health_check():
    return {"status": "OK", "message": "API OTech-MXD funcionando correctamente"}
